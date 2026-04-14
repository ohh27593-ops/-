import gspread
from oauth2client.service_account import ServiceAccountCredentials
import openpyxl
import pyautogui
import time

# ---------------------------
# 1. 구글 시트 인증 및 연결
# ---------------------------
scope = ["https://spreadsheets.google.com/feeds",
         "https://www.googleapis.com/auth/spreadsheets",
         "https://www.googleapis.com/auth/drive"]
creds = ServiceAccountCredentials.from_json_keyfile_name("credentials.json", scope)
client = gspread.authorize(creds)

# 구글 시트 열기
spreadsheet = client.open_by_key("your_sheet_id")
worksheet = spreadsheet.worksheet("sheet_name")
sheet_data = worksheet.get_all_records()  # 리스트[딕셔너리]

# ---------------------------
# 2. 로컬 Excel 파일 열기
# ---------------------------
excel_path = r"data/contact_list.xlsx"
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

# 기존 전화번호 목록 추출
existing_phones = set()
for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
    phone = str(row[1])  # B열: 전화번호
    existing_phones.add(phone)

# ---------------------------
# 3. 신규 인원만 추가
# ---------------------------
new_entries = []  # 새로 추가된 사람 저장용
for row in sheet_data:
    name = row['이름']
    phone = str(row['전화번호'])

    if phone not in existing_phones:
        ws.append([name, phone, ''])  # 전송여부는 일단 빈칸
        existing_phones.add(phone)
        new_entries.append((name, phone))  # 나중에 카톡 보낼 사람들 저장

# ---------------------------
# 4. 카카오톡 메시지 자동 전송
# ---------------------------
for idx, (name, phone) in enumerate(new_entries):
    message = f"""[복돌복실] 안녕하세요 :)
신청 감사드립니다!
본인 확인을 위해, 구글폼에 적은 '이름'을 이 메시지에 회신해주세요."""

    # 카카오톡 친구 검색 단축키 (Ctrl+Alt+F)
    pyautogui.hotkey('ctrl', 'alt', 'f')
    time.sleep(0.5)

    # 전화번호로 검색 → Enter
    pyautogui.write(phone)
    pyautogui.press('enter')
    time.sleep(0.5)

    # 메시지 전송
    pyautogui.write(message)
    pyautogui.press('enter')
    print(f"{name} ({phone}) 에게 카톡 전송 완료.")

    # Excel에서 해당 행의 '전송여부'에 'O' 기록
    ws.cell(row=ws.max_row - len(new_entries) + idx + 1, column=3).value = 'O'  # C열

    time.sleep(1)  # 다음 사람 전송까지 딜레이

# ---------------------------
# 5. Excel 저장
# ---------------------------
wb.save(excel_path)
print(f"총 {len(new_entries)}명에게 카카오톡 전송 및 Excel 기록 완료!")
